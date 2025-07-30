import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_structure():
    """Analyze the Excel file structure to understand named ranges and data"""
    
    # Load the workbook
    wb = load_workbook('Bauphase.xlsm', data_only=True)
    
    print("Available sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    print("\nNamed ranges:")
    try:
        for name in wb.defined_names:
            print(f"  - {name}: {wb.defined_names[name]}")
    except Exception as e:
        print(f"  Error accessing named ranges: {e}")
    
    # Check Schedule sheet for ID list
    if 'Schedule' in wb.sheetnames:
        ws = wb['Schedule']
        print(f"\nSchedule sheet dimensions: {ws.dimensions}")
        
        # Look for cells with values that might be IDs
        print("\nNon-empty cells in Schedule sheet:")
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
            for cell in row:
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {cell.value}")
    
    # Check Template sheet
    if 'Template' in wb.sheetnames:
        ws = wb['Template']
        print(f"\nTemplate sheet dimensions: {ws.dimensions}")
        
        # Look for named cells
        print("\nNon-empty cells in Template sheet:")
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
            for cell in row:
                if cell.value is not None:
                    print(f"  {cell.coordinate}: {cell.value}")

if __name__ == "__main__":
    analyze_excel_structure() 