import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import win32com.client
import re

class SimpleExcelProcessor:
    """Simplified Python equivalent of the VBA code"""
    
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.wb = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.wb = load_workbook(self.excel_file_path, data_only=True)
            print(f"Loaded workbook: {self.excel_file_path}")
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def get_id_list(self):
        """Get list of valid IDs from Schedule sheet"""
        if 'Schedule' not in self.wb.sheetnames:
            print("Schedule sheet not found")
            return []
        
        ws = self.wb['Schedule']
        ids = []
        
        # Look for ID patterns in the sheet
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10):
            for cell in row:
                if cell.value is not None:
                    value = str(cell.value).strip()
                    # Look for lighting component IDs
                    if any(prefix in value for prefix in ['LC-', 'LW-', 'LT-', 'LJ-']):
                        # Clean invalid characters
                        clean_value = re.sub(r'[\[\]*?/\\:;]', '', value).strip()
                        if clean_value and len(clean_value) <= 31:
                            ids.append(clean_value)
                            print(f"Found ID: {clean_value}")
        
        return sorted(list(set(ids)))
    
    def create_sheets(self):
        """Create sheets for each ID"""
        if 'Template' not in self.wb.sheetnames:
            print("Template sheet not found")
            return False
        
        ids = self.get_id_list()
        if not ids:
            print("No valid IDs found")
            return False
        
        template_sheet = self.wb['Template']
        
        for sheet_id in ids:
            if sheet_id not in self.wb.sheetnames:
                print(f"Creating sheet: {sheet_id}")
                
                # Copy template sheet
                new_sheet = self.wb.copy_worksheet(template_sheet)
                new_sheet.title = sheet_id
                
                # Try to set selection cell
                try:
                    # Look for cells that might be selection cells
                    for row in new_sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
                        for cell in row:
                            if cell.value and any(pattern in str(cell.value) for pattern in ['LC-', 'LW-', 'LT-', 'LJ-']):
                                cell.value = sheet_id
                                print(f"Set selection cell to: {sheet_id}")
                                break
                        else:
                            continue
                        break
                except Exception as e:
                    print(f"Warning: Could not set selection cell for {sheet_id}")
            else:
                print(f"Sheet {sheet_id} already exists")
        
        # Save workbook
        try:
            self.wb.save(self.excel_file_path)
            print("Workbook saved successfully")
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def create_pdf(self):
        """Create PDF from all sheets"""
        output_pdf = os.path.splitext(self.excel_file_path)[0] + "_output.pdf"
        print(f"Creating PDF: {output_pdf}")
        
        try:
            # Use Excel automation
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            # Open workbook
            workbook = excel_app.Workbooks.Open(os.path.abspath(self.excel_file_path))
            
            # Export to PDF
            workbook.ExportAsFixedFormat(
                Type=0,  # PDF
                Filename=os.path.abspath(output_pdf),
                Quality=0,
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            # Clean up
            workbook.Close(SaveChanges=False)
            excel_app.Quit()
            
            print(f"PDF created: {output_pdf}")
            return True
            
        except Exception as e:
            print(f"Error creating PDF: {e}")
            try:
                if 'workbook' in locals():
                    workbook.Close(SaveChanges=False)
                if 'excel_app' in locals():
                    excel_app.Quit()
            except:
                pass
            return False
    
    def process(self):
        """Main processing function"""
        print("Starting Excel processing...")
        
        # Load workbook
        if not self.load_workbook():
            return False
        
        # Create sheets
        if not self.create_sheets():
            return False
        
        # Create PDF
        if not self.create_pdf():
            return False
        
        print("Processing completed!")
        return True

def main():
    excel_file = "Bauphase.xlsm"
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found")
        return
    
    processor = SimpleExcelProcessor(excel_file)
    processor.process()

if __name__ == "__main__":
    main() 