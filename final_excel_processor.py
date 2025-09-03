from typing import List
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import sys
import win32com.client
import re
import shutil
from datetime import datetime

def create_backup(excel_file_path):
    """Create a backup copy of the original file"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"{os.path.splitext(excel_file_path)[0]}_backup_{timestamp}.xlsx"
    
    try:
        shutil.copy2(excel_file_path, backup_path)
        print(f"Created backup: {backup_path}")
        return backup_path
    except Exception as e:
        print(f"Warning: Could not create backup: {e}")
        return None


def add_image_to_sheet(sheet, sheet_id, img_dir):
    """Add image to sheet based on sheet ID"""
    try:
        image_path = os.path.join(img_dir, f"{sheet_id}_image.jpg")
        
        # Check if image exists
        if not os.path.exists(image_path):
            print(f"Warning: Image not found for {sheet_id}: {image_path}")
            # Try alternative image paths
            alternative_paths = [
                os.path.join(img_dir, f"{sheet_id}_dimensions.jpg"),
                os.path.join(img_dir, "_no_image.jpg"),
                os.path.join(img_dir, "_blank.jpg")
            ]
            
            for alt_path in alternative_paths:
                if os.path.exists(alt_path):
                    image_path = alt_path
                    print(f"Using alternative image: {alt_path}")
                    break
            else:
                print(f"No suitable image found for {sheet_id}")
                return False
        
        # Load and insert image
        img = Image(image_path)
        
        # Resize image while maintaining aspect ratio
        # Set maximum dimensions (adjust these values as needed)
        max_width = 300
        max_height = 200
        
        # Calculate new dimensions while maintaining aspect ratio
        original_width = img.width
        original_height = img.height
        
        # Calculate scaling factors
        width_ratio = max_width / original_width
        height_ratio = max_height / original_height
        
        # Use the smaller ratio to ensure image fits within bounds
        scale_factor = min(width_ratio, height_ratio)
        
        # Only resize if the image is larger than max dimensions
        if scale_factor < 1:
            img.width = int(original_width * scale_factor)
            img.height = int(original_height * scale_factor)
            print(f"Resized image from {original_width}x{original_height} to {img.width}x{img.height}")
        else:
            print(f"Image size {original_width}x{original_height} is within limits, no resizing needed")
        
        # Insert image at a specific cell (adjust cell position as needed)
        # Common positions: 'A1', 'B1', 'A2', etc.
        sheet.add_image(img, 'D15')
        
        print(f"Added image for {sheet_id}: {image_path}")
        return True
        
    except Exception as e:
        print(f"Error adding image for {sheet_id}: {e}")
        return False

def create_sheets(wb, excel_file_path, language, img_dir):
    """Create sheets directly from Schedule sheet data using openpyxl"""
    template_sheet_name = f'Template_{language}'
    
    if template_sheet_name not in wb.sheetnames:
        print("Template sheet not found")
        return False
    
    if 'Schedule' not in wb.sheetnames:
        print("Schedule sheet not found")
        return False
    
    try:
        # Get the template sheet
        template_sheet = wb[template_sheet_name]
        
        # Get the Schedule sheet
        schedule_sheet = wb['Schedule']
        sheets_created = 0
        
        # Get column names from row 9
        column_names = {}
        for col_num in range(1, schedule_sheet.max_column + 1):
            cell_value = schedule_sheet.cell(row=9, column=col_num).value
            if cell_value:
                column_names[col_num] = str(cell_value).strip()
        
        print(f"Found columns: {list(column_names.values())}")
        
        # Process each row in the Schedule sheet starting from row 11
        sheet_ids = []
        for row_num in range(11, schedule_sheet.max_row + 1):
            # Get all values for this row
            row_data = {}
            for col_num in column_names.keys():
                cell_value = schedule_sheet.cell(row=row_num, column=col_num).value
                row_data[column_names[col_num]] = cell_value
            
            cell_value = row_data.get("ID")
            
            if cell_value is None or str(cell_value).strip() == '':
                # Stop when we encounter a blank value
                break
            
            value_str = str(cell_value).strip()
            # Clean invalid characters
            sheet_id = re.sub(r'[\[\]*?/\\:;]', '', value_str).strip()
            sheet_ids.append(sheet_id)
            if sheet_id and len(sheet_id) <= 31 and sheet_id not in wb.sheetnames:
                print(f"Creating sheet: {sheet_id}")
                print(f"Row data: {row_data}")
                
                # Copy the template sheet
                new_sheet = wb.copy_worksheet(template_sheet)
                new_sheet.title = sheet_id
                
                # Try to update cells that contain lighting component patterns
                for row in new_sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
                    for cell in row:
                        if cell.value and any(pattern in str(cell.value) for pattern in ['LC-', 'LW-', 'LT-', 'LJ-']):
                            cell.value = sheet_id
                            print(f"Set selection cell to: {sheet_id}")
                            break
                    else:
                        continue
                    break
                
                # Add image to the sheet
                add_image_to_sheet(new_sheet, sheet_id, img_dir)
                
                sheets_created += 1
            elif sheet_id in wb.sheetnames:
                print(f"Sheet {sheet_id} already exists, deleting and recreating...")
                # Remove the existing sheet
                wb.remove(wb[sheet_id])
                print(f"Deleted existing sheet: {sheet_id}")
                
                # Create the new sheet (this will now execute the code below)
                print(f"Creating sheet: {sheet_id}")
                print(f"Row data: {row_data}")
                
                # Copy the template sheet
                new_sheet = wb.copy_worksheet(template_sheet)
                new_sheet.title = sheet_id
                
                # Try to update cells that contain lighting component patterns
                for row in new_sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
                    for cell in row:
                        if cell.value and any(pattern in str(cell.value) for pattern in ['LC-', 'LW-', 'LT-', 'LJ-']):
                            cell.value = sheet_id
                            print(f"Set selection cell to: {sheet_id}")
                            break
                    else:
                        continue
                    break
                
                # Add image to the sheet
                add_image_to_sheet(new_sheet, sheet_id, img_dir)
                
                sheets_created += 1
        
        # Save the workbook
        wb.save(excel_file_path)
        print(f"Created {sheets_created} new sheets")
        print("Workbook saved successfully")
        return sheet_ids
        
    except Exception as e:
        print(f"Error creating sheets: {e}")
        print("Trying to save with different name...")
        
        # Try saving with a different name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_path = f"{os.path.splitext(excel_file_path)[0]}_modified_{timestamp}.xlsm"
        
        try:
            wb.save(new_path)
            print(f"Workbook saved as: {new_path}")
            return new_path
        except Exception as e2:
            print(f"Error saving to new path: {e2}")
            return False

def create_pdf(excel_file_path, sheet_ids: List[str]):
    """Create PDF from all sheets"""
    output_pdf = os.path.splitext(excel_file_path)[0] + "_output.pdf"
    if os.path.exists(output_pdf):
        os.remove(output_pdf)
    print(f"Creating PDF: {output_pdf}")
    
    try:
        # Use Excel automation
        excel_app = win32com.client.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        
        # Open workbook
        workbook = excel_app.Workbooks.Open(os.path.abspath(excel_file_path))

        # Define sheets to include in PDF export
        sheets_to_include = ["Cover", "GenInfo+Contacts"]
        sheets_to_include = sheets_to_include + sheet_ids
        
        print(f"Sheets to include in PDF: {sheets_to_include}")
        
        # Hide sheets that should not be included in PDF
        for sheet in workbook.Sheets:
            if sheet.Name not in sheets_to_include:
                sheet.Visible = False
                print(f"Hidden sheet: {sheet.Name}")
        
        # Export to PDF (only visible sheets will be included)
        workbook.ExportAsFixedFormat(
            Type=0,  # PDF
            Filename=os.path.abspath(output_pdf),
            Quality=0,
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False
        )
        
        # Make all sheets visible again
        for sheet in workbook.Sheets:
            sheet.Visible = True
        
        # Clean up
        workbook.Close(SaveChanges=False)
        excel_app.Quit()
        
        print(f"PDF created successfully: {output_pdf}")
        return output_pdf
        
    except Exception as e:
        print(f"Error creating PDF: {e}")
        try:
            if 'workbook' in locals():
                workbook.Close(SaveChanges=False)
            if 'excel_app' in locals():
                excel_app.Quit()
        except:
            pass
        return False

def process_excel_file(excel_file_path, language, img_dir):
    """Main processing function"""
    print("Starting Excel processing and PDF creation...")
    print("=" * 50)
    
    # Create backup
    backup_path = create_backup(excel_file_path)
    
    # Load workbook
    try:
        wb = load_workbook(excel_file_path)
        print(f"Loaded workbook: {excel_file_path}")
    except Exception as e:
        print(f"Error loading workbook: {e}")

    if wb is None:
        return False
    
    # Create sheets
    sheet_ids = create_sheets(wb, excel_file_path, language, img_dir)
    if not sheet_ids:
        return False
    
    # Create PDF
    pdf_path = create_pdf(excel_file_path, sheet_ids)
    if not pdf_path:
        return False
    
    print("=" * 50)
    print("Processing completed successfully!")
    print(f"Modified Excel file: {excel_file_path}")
    print(f"PDF output: {os.path.splitext(excel_file_path)[0]}_output.pdf")
    if backup_path:
        print(f"Backup created: {backup_path}")
    return pdf_path


if __name__ == "__main__":
    excel_file = r"C:\Users\aelnagar\Downloads\Lighting Computational Development\Lighting Computational Development\Specifications - BAM example\Bauphase.xlsx"
    img_dir = r"C:\Users\aelnagar\Downloads\Lighting Computational Development\Lighting Computational Development\Specifications - BAM example\img"
    #excel_file = r"C:\Users\aelnagar\Downloads\Lighting Computational Development\Bauphase.xlsm"
    language = "EN"
    process_excel_file(excel_file, language, img_dir) 