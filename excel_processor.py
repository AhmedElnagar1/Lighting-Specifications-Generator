import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys
from pathlib import Path
import win32com.client
import time
import re

class ExcelProcessor:
    """Python equivalent of the VBA code for processing Excel files"""
    
    # Constants matching VBA code
    SCHEDULE_SHEET_NAME = "Schedule"
    TEMPLATE_SHEET_NAME = "Template"
    ID_LIST_NAME = "list_ids"
    SELECTION_CELL_NAME = "selection"
    IMG_TOPLEFT_CELL_NAME = "image_topleft"
    IMG_TOPRIGHT_CELL_NAME = "image_topright"
    IMG_BOTTOMLEFT_CELL_NAME = "image_bottomleft"
    IMG_BOTTOMRIGHT_CELL_NAME = "image_bottomright"
    
    def __init__(self, excel_file_path):
        """Initialize with the Excel file path"""
        self.excel_file_path = excel_file_path
        self.workbook_path = os.path.dirname(os.path.abspath(excel_file_path))
        self.image_folder = "img"
        self.wb = None
        self.excel_app = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.wb = load_workbook(self.excel_file_path, data_only=True)
            print(f"Successfully loaded workbook: {self.excel_file_path}")
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def worksheet_exists(self, worksheet_name):
        """Check if worksheet exists (equivalent to VBA WorksheetExists function)"""
        return worksheet_name in self.wb.sheetnames
    
    def is_valid_sheet_name(self, name):
        """Check if a sheet name is valid (no invalid characters)"""
        # Excel sheet names cannot contain: [ ] * ? / \
        invalid_chars = ['[', ']', '*', '?', '/', '\\', ':', ';']
        return not any(char in name for char in invalid_chars) and len(name) <= 31
    
    def get_id_list(self):
        """Get the list of IDs from the Schedule sheet"""
        if self.SCHEDULE_SHEET_NAME not in self.wb.sheetnames:
            print(f"Error: {self.SCHEDULE_SHEET_NAME} sheet not found")
            return []
        
        ws = self.wb[self.SCHEDULE_SHEET_NAME]
        ids = []
        
        # Try to find the named range first
        try:
            if self.ID_LIST_NAME in self.wb.defined_names:
                range_ref = self.wb.defined_names[self.ID_LIST_NAME]
                print(f"Found named range {self.ID_LIST_NAME}: {range_ref}")
                
                # Parse the range reference (e.g., 'Schedule!$A$11:$A$47')
                if '!' in str(range_ref):
                    sheet_name, range_str = str(range_ref).split('!')
                    # Extract values from the range
                    try:
                        for row in ws.iter_rows(range_string=range_str):
                            for cell in row:
                                if cell.value is not None:
                                    value = str(cell.value).strip()
                                    if value and self.is_valid_sheet_name(value):
                                        ids.append(value)
                                        print(f"Found ID from named range: {value}")
                    except Exception as e:
                        print(f"Error parsing named range: {e}")
        except Exception as e:
            print(f"Error accessing named range: {e}")
        
        # Fallback: look for ID values in the sheet
        print("Searching for ID values in Schedule sheet...")
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=10):
            for cell in row:
                if cell.value is not None:
                    value = str(cell.value).strip()
                    # Look for patterns that might be IDs (LC-, LW-, LT-, LJ-)
                    if any(prefix in value for prefix in ['LC-', 'LW-', 'LT-', 'LJ-']):
                        # Clean up the value and check if it's a valid sheet name
                        clean_value = re.sub(r'[\[\]*?/\\:;]', '', value).strip()
                        if clean_value and self.is_valid_sheet_name(clean_value):
                            ids.append(clean_value)
                            print(f"Found ID: {clean_value}")
        
        # Remove duplicates and sort
        return sorted(list(set(ids)))
    
    def create_all_sheets(self):
        """Create sheets for all available IDs (equivalent to VBA CreateAllSheets)"""
        print("Creating sheets for all IDs...")
        
        if self.TEMPLATE_SHEET_NAME not in self.wb.sheetnames:
            print(f"Error: {self.TEMPLATE_SHEET_NAME} sheet not found")
            return False
        
        ids = self.get_id_list()
        if not ids:
            print("No valid IDs found to process")
            return False
        
        template_sheet = self.wb[self.TEMPLATE_SHEET_NAME]
        last_sheet = template_sheet
        
        for sheet_id in ids:
            if not self.worksheet_exists(sheet_id):
                print(f"Creating sheet: {sheet_id}")
                
                try:
                    # Copy template sheet
                    new_sheet = self.wb.copy_worksheet(template_sheet)
                    new_sheet.title = sheet_id
                    
                    # Set the selection cell value
                    try:
                        # Try to find the selection cell by name or position
                        selection_cell = None
                        
                        # Method 1: Try direct cell reference
                        try:
                            selection_cell = new_sheet[self.SELECTION_CELL_NAME]
                        except:
                            pass
                        
                        # Method 2: Look for a cell that might contain the selection value
                        if selection_cell is None:
                            for row in new_sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
                                for cell in row:
                                    if cell.value and "selection" in str(cell.value).lower():
                                        selection_cell = cell
                                        break
                                if selection_cell:
                                    break
                        
                        # Method 3: Look for cells with specific patterns
                        if selection_cell is None:
                            for row in new_sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
                                for cell in row:
                                    if cell.value and any(pattern in str(cell.value) for pattern in ['LC-', 'LW-', 'LT-', 'LJ-']):
                                        selection_cell = cell
                                        break
                                if selection_cell:
                                    break
                        
                        if selection_cell:
                            selection_cell.value = sheet_id
                            print(f"Set selection cell to: {sheet_id}")
                        else:
                            print(f"Warning: Could not find selection cell for {sheet_id}")
                            
                    except Exception as e:
                        print(f"Warning: Could not set selection cell for {sheet_id}: {e}")
                    
                    last_sheet = new_sheet
                    
                except Exception as e:
                    print(f"Error creating sheet {sheet_id}: {e}")
                    continue
            else:
                print(f"Sheet {sheet_id} already exists, skipping...")
        
        # Save the workbook
        try:
            self.wb.save(self.excel_file_path)
            print("Workbook saved successfully")
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def clear_all_sheets(self):
        """Delete all automatically created catalogue sheets (equivalent to VBA ClearAllSheets)"""
        print("Clearing all automatically created sheets...")
        
        ids = self.get_id_list()
        sheets_to_remove = []
        
        for sheet_id in ids:
            if self.worksheet_exists(sheet_id):
                sheets_to_remove.append(sheet_id)
        
        for sheet_name in sheets_to_remove:
            try:
                del self.wb[sheet_name]
                print(f"Removed sheet: {sheet_name}")
            except Exception as e:
                print(f"Error removing sheet {sheet_name}: {e}")
        
        # Save the workbook
        try:
            self.wb.save(self.excel_file_path)
            print("Workbook saved after clearing sheets")
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
    
    def create_pdf(self, output_pdf_path=None):
        """Create PDF from all sheets using Excel automation"""
        if output_pdf_path is None:
            output_pdf_path = os.path.splitext(self.excel_file_path)[0] + "_output.pdf"
        
        print(f"Creating PDF: {output_pdf_path}")
        
        try:
            # Use Excel automation to create PDF
            excel_app = win32com.client.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            # Open the workbook
            workbook = excel_app.Workbooks.Open(os.path.abspath(self.excel_file_path))
            
            # Get all sheet names
            sheet_names = [sheet.Name for sheet in workbook.Sheets]
            print(f"Sheets to include in PDF: {sheet_names}")
            
            # Export to PDF
            workbook.ExportAsFixedFormat(
                Type=0,  # PDF
                Filename=os.path.abspath(output_pdf_path),
                Quality=0,  # Standard quality
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False
            )
            
            # Close workbook and quit Excel
            workbook.Close(SaveChanges=False)
            excel_app.Quit()
            
            print(f"PDF created successfully: {output_pdf_path}")
            return True
            
        except Exception as e:
            print(f"Error creating PDF: {e}")
            try:
                if 'workbook' in locals():
                    workbook.Close(SaveChanges=False)
                if 'excel_app' in locals():
                    excel_app.Quit()
            except:
                pass
            return False
    
    def process_and_create_pdf(self):
        """Main function to process Excel file and create PDF"""
        print("Starting Excel processing and PDF creation...")
        
        # Load workbook
        if not self.load_workbook():
            return False
        
        # Create sheets for all IDs
        if not self.create_all_sheets():
            return False
        
        # Create PDF
        if not self.create_pdf():
            return False
        
        print("Processing completed successfully!")
        return True

def main():
    """Main function to run the Excel processor"""
    excel_file = "Bauphase.xlsm"
    
    if not os.path.exists(excel_file):
        print(f"Error: {excel_file} not found in current directory")
        return
    
    processor = ExcelProcessor(excel_file)
    processor.process_and_create_pdf()

if __name__ == "__main__":
    main() 